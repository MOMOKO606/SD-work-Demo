from openpyxl import Workbook  # 写入excel用的库。
import openpyxl.styles as sty  # 修改excel颜色。

wb = Workbook()  # 在内存中创建一个workbook对象，而自动创建一个worksheet。
ws = wb.active  # 获取当前活跃的worksheet，默认就是第一个worksheet。

ws.append([1,2,3,4])  # worksheet中写入一行光敏数据。
ws.cell(row=1,column=1).fill=sty.PatternFill(fill_type='solid',fgColor="007A378B")
#ws.cell(row = 1,column = 1).fill =sty.PatternFill(fill_type = "solid", fgColor = "#FFB6C1")

wb.save("test.xlsx")


