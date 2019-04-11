import time
import math
import numpy as np
from openpyxl import Workbook  # 写入excel用的库。
import openpyxl.styles as sty  # 修改excel颜色。


class PsTrackers( list ):
    """
    定义16个光敏追踪器的类。
    """
    def __init__(self, timewindow):
        """
        function: 初始化16个光敏追踪器，即(timewindow*6) * 16的空table, 可视为是16个队列组成的table。
        :param timewindow: 追踪多长时间的光敏值，单位为分钟，至少1分钟，建议3分钟。
        """

        #  Sentinel.
        assert timewindow >= 1, "光敏追踪器时间窗至少为1分钟。"

        list.__init__([])  # 继承list类。
        m = timewindow * 6 + 1  # 光敏追踪器的行数, 加1是由于队列的结构。
        n = 16
        self.trackers = [[[] for i in range(n)] for j in range(m)]  # 追踪器初值。
        self.brokelist = [False for i in range(16)]  # 异常1的光敏列表。
        self.brokegratings = [False for i in range(16)]  # 异常2的光敏列表。
        self.gratingcount = [0.0 for i in range(16)]  # 判断异常1的辅助数组。
        self.sum = [0.0 for i in range(16)]  # 追踪器的总值。
        self.mean = [0.0 for i in range(16)]  # 追踪器的均值。
        self.stddev = [0.0 for i in range(16)]  # 追踪器的标准差。
        self.rows = m  # 由timewindow转换的行数。
        self.head = 0  # 光敏追踪器的head。
        self.tail = 0  # 光敏追踪器的tail。
        self.count = 0  # 表示追踪器内目前存有多少条数据。



    def reset(self):
        """
        function: 清除追踪器内的所有数据。
        """
        n = 16
        m = self.rows
        self.trackers = [[[] for i in range(n)] for j in range(m)]  # 追踪器初值。
        self.brokelist = [False for i in range(16)]  #  异常1光敏列表。
        self.brokegratings = [False for i in range(16)]  # 异常2的光敏列表。
        self.gratingcount = [0.0 for i in range(16)]  # 判断异常1的辅助数组。
        self.mean = [0 for i in range(16)]   # 追踪器的均值。
        self.sum = [0 for i in range(16)]  # 追踪器的总值。
        self.stddev = [0 for i in range(16)]   #  追踪器的标准差。
        self.head = 0  # 光敏追踪器的head。
        self.tail = 0  # 光敏追踪器的tail。
        self.count = 0  #  表示追踪器内有多少条数据。



    def isempty( self ):
        """
        function:判断当前队列是否为空。
        :return: True if it is empty, False otherwise.
        """
        if self.tail == self.head:
            return True
        else:
            return False



    def isfull( self ):
        """
        function:判断当前队列是否已满。
        :return: True if it is full, False otherwise.
        """
        if (self.tail + 1 == self.head) or (self.head == 0 and self.tail == self.rows - 1):
            return True
        else:
            return False



    def dequeue( self ):
        """
        function:从队列table的head弹出一行元素。
        :return: 弹出的一行16个元素值。
        """

        #  取出参数。
        trackers = self.trackers
        head = self.head
        tail = self.tail
        rows = self.rows

        #  Sentinel.
        assert not self.isempty(), "Empty queue!"

        #  弹出的数据。
        res = [trackers[head][j] for j in range(16)]
        trackers[head] = [[] for j in range(16)]

        #  检测弹出的数据是属于异常1的光敏。
        for j in range(16):
            if self.brokelist[j]:
                #  因为是弹出该数据，所以对应gratingcount减1。
                self.gratingcount[j] -= 1
        #  修改head的位置。
        if head == rows - 1:
            head = 0
        else:
            head += 1

        #  更新属性。
        self.head = head
        self.trackers = trackers
        return res



    def GetIndices(self):
        """
        function: 得到从tail向head方向的角标indices.
        :return: indices数组。
        """
        #  读取参数。
        head = self.head
        tail = self.tail
        rows = self.rows

        #  case1.queue第一次插满后，又插入一条数据，tail回到0角标的情况。
        if tail == 0:
            indices = [rows - 1, 0]
        #  case2.queue不满时。
        elif tail > head:
            indices = [tail - 1, head - 1]
        #  case3.head > tail，除case1外，queue插满时还在继续插入数据的情况。
        else:
            indices = [[tail - 1, -1], [rows - 1, head - 1]]
        return indices



    def Isbroken(self):
        """
        function: detecting each photosensitive to see whether is good or broken.

        异常1.光敏完全损坏。
        检测：检查追踪器中的标准差。如果光敏损坏，会产生0到1023的随机值，其随机分布标准差在295左右。
             所以当追踪器中的标准差接近250时，说明该光敏值是随机的，即光敏损坏。
             但是，如果泥位迅速增高或下降时，也有可能产生0，200，400，600，... , 900的情况。
             这种情况的标准差较大，但其两两光敏值的差值是同一方向的，使用分位数判断是否属于该情况。
        return: 完全损坏的光敏列表brokelist。
        """

        #  载入数据。
        rows = self.rows
        brokelist = self.brokelist
        trackers = self.trackers

        #  定义常量
        STANDARD_STDDEV = 220  # 0-1023随机分布的标准差约等于295。

        #  首先要知道从近到远的历史数据顺序，
        #  即通过GetIndices函数得到从tail到head的角标。
        indices = self.GetIndices()
        #  对应GetIndices函数中的case1和case2.
        if len(np.array(indices).shape) == 1:
            start = indices[0]
            end = indices[1]
            delta = ( start - end ) // 5

            #  检查光敏是否完全损坏。
            #  1分钟以上检测(长周期检测)。
            for j in range(16):
                #  如果被测为异常1，则不再重测。
                #  设备维修后请reset光敏追踪器。
                if self.brokelist[j]:
                    continue
                #  标准差大于STANDARD_STDDEV,波动较大，疑似随机值：
                if self.stddev[j] > STANDARD_STDDEV:
                    #  检查时间窗内，追踪器分位数数据是否单调递增或递减。
                    #  单调递增，则由泥变水；
                    #  单调递减，则由水变泥。
                    if start - end >= 5:
                        flag = 1
                        for i in range(start, end + delta, -delta):
                            flag *= (trackers[i][j] - trackers[i - delta][j])
                            #  不是单调的，说明是随机值，即光敏损坏。
                        if flag < 0:
                            self.gratingcount[j] += 1
                    else:
                        self.gratingcount[j] += 1
        #  对应GetIndices函数中的case3.
        else:
            start1 = indices[0][0]
            end1 = indices[0][1]
            start2 = indices[1][0]
            end2 = indices[1][1]
            delta = start1 - end1 + 1 + start2 - end2 + 1

            #  1分钟以上检测(长周期检测)。
            for j in range(16):
                #  如果被测为异常1，则不再重测。
                #  设备维修后请reset光敏追踪器。
                if self.brokelist[j]:
                    continue
                #  标准差大于STANDARD_STDDEV,波动较大，疑似随机值：
                if self.stddev[j] > STANDARD_STDDEV:
                    #  检查时间窗内，追踪器数据是否单调递增或递减。
                    #  单调递增，则由泥变水；
                    #  单调递减，则由水变泥。
                    flag = 1
                    for i in range(start1, end1 + delta - 1, -delta):
                        flag *= (trackers[i][j] - trackers[i - delta][j])
                    for i in range(start2, end2 + delta - 1, -delta):
                        flag *= (trackers[i][j] - trackers[i - delta][j])
                    #  不是单调的，说明是随机值，即光敏损坏。
                    if flag < 0:
                        self.gratingcount[j] += 1

        #  更新仪器损坏表。
        for i in range(16):
            if self.gratingcount[i] / self.count >= 1 / 2:
                brokelist[i] = True
            # 自动检测光敏修复效果不好。
            # else:
            #     brokelist[i] = False
        self.brokelist = brokelist
        return brokelist



    def AddBound(self, psvalues):
        """
        function: 对16个光敏数据的首位添加边界，最左侧添加0，最右侧添加900。
        :return: 添加边界后的16+2个光敏数据。
        """

        n = len(psvalues)
        res = [0.0 for i in range(0, n + 2)]
        res[n + 1] = 900
        for i in range(1, n + 1):
            res[i] = psvalues[i - 1]
        return res



    def L2Rmoving(self, start, end, leftkey, A, res):
        """
        function: 从start向end前进，并鉴别每个数据是否发生异常。
        :param start: 输入数据A的起始下标。
        :param end: 输入数据A的终点下标。
        :param leftkey: start左侧，即上一阶段的最后一个正常值。
        :param A: 输入数据。
        :param res: 对应输入数据A的bool数组，False表示该点数据正常，True表示该点数据存在异常2的情况。
                    异常2包括：各别光栅损坏，浮泥，泥中含较多的水（简称含水泥）。
        :return:
        """

        GRATING_DELTA = 200 # 光敏值敏感间隔。
        #  base case: 已遍历完。
        if start > end:
            return res

        #  循环初值。
        i = start
        while True:
            #  未到达边界时，i或i + 1处光敏发生异常1, 右移。
            #  注意，A[0, ..., 17]加过左右边界的，brokelist[0,...,15]的，所以是i - 1。
            if i < 16:
                if self.brokelist[i - 1] is True or self.brokelist[i] is True:
                    i += 1
                    continue
            #  到达边界时，且i处光敏发生异常1，跳出循环。
            else:
                if self.brokelist[i - 1] is True:
                    break

            #  i和i + 1处光敏没有发生异常1。
            rightdelta = A[i] - A[i + 1]
            #  右侧数据与当前数据相差不大, 且没有到右边界，则右移。
            if abs(rightdelta) <= GRATING_DELTA and i < end:
                i += 1
            #  否则跳出循环。
            else:
                break

        #  右侧数据发生跳跃，鉴别该阶段数据是否正常。
        if abs(rightdelta) > GRATING_DELTA:
            leftdelta = A[i] - leftkey
            #  发生相变的正常情况。
            if abs(leftdelta) <= GRATING_DELTA or leftdelta * rightdelta < 0:
                for j in range(start, i + 1):
                    res[j] = False
                leftkey = A[i]
            #  可能发生异常2。
            else:
                if i - start + 1 < 5:
                    #   排除“900, 900, 900, 30”最高位浮泥的情况。
                    if i + 1 == end and A[end] < 300:
                        for j in range(start, i + 1):
                            res[j] = False
                        leftkey = A[i]
                    #  发生异常2，leftkey不变。
                    else:
                        for j in range(start, i + 1):
                            res[j] = True
                #  排除连坏5个的情况。
                else:
                    for j in range(start, i + 1):
                        res[j] = False
                    leftkey = A[i]
        #  向右跳转到下一阶段。
        return self.L2Rmoving(i + 1, end, leftkey, A, res)



    def enqueue(self, psvalues):
        """
        function:对光敏追踪器增加一行数据psvalues，并更新各种属性。
        :param psvalues: 实时光敏数据。
        """

        #  取出参数。
        trackers = self.trackers
        rows = self.rows
        head = self.head
        tail = self.tail

        #  更新count, sum, mean属性。
        #  如果队列已满，则先从head弹出一行数据。
        popedvalue = []
        if self.isfull():
            popedvalue = self.dequeue()
            self.count = self.rows - 1
            self.sum = [self.sum[i]-popedvalue[i] + psvalues[i] for i in range(16)]
        #  队列未存满时：
        else:
            self.count += 1
            self.sum = [self.sum[i] + psvalues[i] for i in range(16)]
        self.mean = [self.sum[i] / self.count for i in range(16)]

        #  在tail处插入数据。
        trackers[tail] = [psvalues[i] for i in range(16)]
        self.trackers = trackers  # 更新属性。
        #  修改tail的值。
        if tail == rows - 1:
            tail = 0
        else:
            tail += 1
        self.tail = tail  # 更新属性。

        #  更新stddev。
        #  遍历追踪器。
        for j in range(16):
            tmp = 0  # 累加器，初值为0。
            for i in range(0, rows):
                #  Sentinel, 跳过空值。
                if trackers[i][j] == []:
                    continue
                tmp += (trackers[i][j] - self.mean[j]) ** 2
            tmp /= self.count
            self.stddev[j] = math.sqrt(tmp)

        #  检测插入的数据是否包含异常1 & 2。
        #  异常1包括：光敏彻底损坏，数值在0到1023之间随机跳动。
        #  检查并更新数据异常1的情况。
        self.brokelist = self.Isbroken()
        #  异常2包括：各别光栅损坏，浮泥，泥中含较多的水（简称含水泥）。
        #  检查并更新数据异常2的情况。
        #  首先对16个光敏数据的首尾添加边界。
        psvtmp = self.AddBound(psvalues)
        #  其次构造1*18的bool数组，存放异常2的检测结果。
        abnor_tmp = [False for i in range(18)]
        #  遍历1*18个光敏数据，检查是否有异常2情况。
        abnor_tmp = self.L2Rmoving(1, 16, 0, psvtmp, abnor_tmp)
        #  去掉首尾边界。
        for i in range(1, 17):
            self.brokegratings[i - 1] = abnor_tmp[i]
        return popedvalue  # 返回弹出的一行数据。



def FuzzyInterval(brokelist, brokegratings, psvalues):
    """
    function:根据当前的16个光敏值判断泥面位于的区间。
    :param brokelist: 当前损坏的光敏列表。
    :param brokegratings: 光栅存在问题的光敏列表。
    :param psvalues: 当前的光敏值。
    :return: 泥位位于的光敏区间。
    """
    #  规则：
    #  （1）数据由左向右表示由低到高，即i = 0表示最低处的光敏，i = 15表示最高处的光敏；
    #  （2）光敏值小于50为泥, [50,200]为泥水混合物, 大于200为水；
    #  （3）光敏有损坏则右移一位。

    #  从左向右确定纯泥界面:
    l = float("-inf")
    for i in range(16):
        #  如果i处的光敏有异常，则右移动一位。
        if brokelist[i] or brokegratings[i]:
            continue
        if psvalues[i] <= 50:
            if i > l:
                l = i
    #  所有光敏值都大于50时，即泥位最低的情况。
    if l == float("-inf"):
        l = 0

    #  从右向左确定纯泥界面:
    r = float("inf")
    for i in range(15, -1, -1):
        #  如果i处的光敏有异常，则左移动一位。
        if brokelist[i] or brokegratings[i]:
            continue
        if psvalues[i] >= 200:
            if i < r:
                r = i
    #  r位是最左侧的大于200的光敏，则r左侧第一个正常值为泥界面。
    #  所有光敏值都小于200时，即泥位最高的情况。
    if r == float("inf"):
        r = 15
    elif r != 0:
        for j in range(r - 1, -1, -1):
            if brokelist[j] or brokegratings[j]:
                continue
            else:
                r = j
                break
    #  [l, r]组成泥位区间。
    if l > r:
        return [r, l]
    #  当l小于等于r时，[l, r]即为泥位区间。
    else:
        return [l, r]



def GetMudLvl( inputfile ):
    """
    function: 根据光敏追踪器类，预处理后的历史光敏数据，计算从泥到水的区间。
    :param inputfile: 预处理后的历史光敏数据，文本文件。
    :return: 显示全部历史光敏数据的excel文件，用红色标出泥位，绿色为异常2的光敏，紫色为异常1的光敏。
    """

    #  新建并初始化光敏追踪器。
    timewindow = 6  # 设定追踪器时间窗长，单位为分钟，最小值为1，建议选取3-6。
    pst = PsTrackers( timewindow )

    #  打开一个excel文件， 用于写入所有的光敏数据。
    wb = Workbook()  # 在内存中创建一个workbook对象，而自动创建一个worksheet。
    ws = wb.active  # 获取当前活跃的worksheet，默认就是第一个worksheet。

    with open( inputfile ) as f:
        ###  测试用  ###
        timecount = 0
        ###  测试用  ###
        while True:
            ###  测试用  ###
            timecount += 1
            ###  测试用  ###
            # 读入光敏数据。
            line = f.readline()  # 读取一行光敏数据（str）。
            #  EOF Sentinel.
            if line == '':
                break
            line = line.strip('\n')  # 删除换行符。
            line = line.split(",")  # 按逗号分割字符串。
            #  将str转换为float，PhotoSensitive Value。
            #  "000"表示1000。
            psvalues = [float(str) if str != "000" else 1000 for str in line]
            #  Collecting the photosensitive values.
            pst.enqueue(psvalues)
            #  异常1 & 2的光敏。
            brokelist = pst.brokelist
            brokegratings = pst.brokegratings
            #  计算泥位。
            mudinterval = FuzzyInterval(brokelist, brokegratings, psvalues)

            print(timecount, mudinterval)

            #  光敏数据写入excel，用于人工检查。
            ws.append(psvalues)  # worksheet中写入一行光敏数据。
            for j in range(16):
                #  将异常2光敏标记成绿色。
                if brokegratings[j]:
                    ws.cell(row = timecount, column = j + 1).fill = \
                        sty.PatternFill(fill_type = "solid", fgColor = "0000FF00")
                #  将异常1光敏标记成紫色。
                if brokelist[j]:
                    ws.cell(row = timecount, column = j + 1).fill = \
                        sty.PatternFill(fill_type = "solid", fgColor = "007A378B")
            #  将泥位标记为红色。
            ws.cell(row=timecount, column=mudinterval[1] + 1).fill = \
                sty.PatternFill(fill_type="solid", fgColor="00FF0000")

        wb.save(inputfile[:-3] + "xlsx")  # 保存为与输入数据文件同名的excel文件。



"""
脚本功能：根据历史数据库中的光敏数据，定位随时间变化的泥位点。

Step0.预处理光敏数据：
      (1)选出一个池子，某段时间内的16组光敏数据，保存为txt文档.
      e.g. 选用db_181013/sanchang_db数据中的5号池，共262943个数（大约1个月），默认命名为PreData02.txt。
Step1.定义光敏追踪器类，纵向追踪历史数据，判断是否属于异常1。
Step2.过滤异常1数据，横向判断是否属于异常2.
Step3.过滤异常1 & 2数据，再横向计算泥位。

注
异常1包括：光敏彻底损坏，数值在0到1023之间随机跳动。
异常2包括：各别光栅损坏，浮泥，泥中含较多的水（简称含水泥），数值表现为大值中含有小值，小值中含有大值。
"""

start = time.time()  # 程序开始时间。

#  预处理后的光敏数据文件名。
inputfile = "fac5_pool12_02.txt"

#  调用函数计算泥位。
#  输出excel文件，用颜色标识泥位区间，损坏的光敏。
GetMudLvl( inputfile )

end = time.time()  # 程序结束时间。
#  显示程序的运行时间，单位为秒。
print(end-start)


